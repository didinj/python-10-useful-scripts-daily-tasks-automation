from openpyxl import load_workbook

def update_excel(file_path, sheet_name, cell, value):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    ws[cell] = value
    wb.save(file_path)
    print(f"📊 Updated {sheet_name}:{cell} with value '{value}'")

# Example usage
# update_excel("report.xlsx", "Sheet1", "B2", 1500)
